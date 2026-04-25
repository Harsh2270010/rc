import time
import io
import json
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────
BASE_URL = "https://api.subhxcosmo.in/api"
API_KEY = "DEMO6"
API_TYPE = "vehicle"
RETRIES = 2

# ─── ALL RESPONSE KEYS (COMPLETE) ─────────────────────────────────────────
FULL_RESPONSE_KEYS = {
    # Top level
    "registration_number": "result.registration_number",
    "success": "success",
    "cached": "cached",
    "proxy_used": "proxyUsed",
    "attempt": "attempt",

    # vnum_chassis_response (58 fields)
    "vnum_chassis_no": "result.vnum_chassis_response.chassis_no",
    "vnum_date_of_purchase": "result.vnum_chassis_response.date_of_purchase",
    "vnum_engine_no": "result.vnum_chassis_response.engine_no",
    "vnum_ex_showroom_price": "result.vnum_chassis_response.ex_showroom_price",
    "vnum_fuel_type": "result.vnum_chassis_response.fuel_type",
    "vnum_make_id": "result.vnum_chassis_response.make_id",
    "vnum_manufacture_month": "result.vnum_chassis_response.manufacture_month",
    "vnum_manufacture_year": "result.vnum_chassis_response.manufacture_year",
    "vnum_model_id": "result.vnum_chassis_response.model_id",
    "vnum_registration_date": "result.vnum_chassis_response.registration_date",
    "vnum_registration_number": "result.vnum_chassis_response.registration_number",
    "vnum_rto_location_id": "result.vnum_chassis_response.rto_location_id",
    "vnum_rto_name": "result.vnum_chassis_response.rto_name",
    "vnum_seating_capacity": "result.vnum_chassis_response.seating_capacity",
    "vnum_state_of_registration_id": "result.vnum_chassis_response.state_of_registration_id",
    "vnum_success": "result.vnum_chassis_response.success",
    "vnum_variant": "result.vnum_chassis_response.variant",
    "vnum_vehicle_name": "result.vnum_chassis_response.vehicle_name",

    # chola2x_response - basic info
    "chola_success": "result.chola2x_response.success",
    "chola_vehicle": "result.chola2x_response.vehicle",
    "chola_source": "result.chola2x_response.source",
    "chola_vehicle_type": "result.chola2x_response.vehicle_type",
    "chola_injected_ip": "result.chola2x_response.injected_ip",

    # chola2x_response.data - vehicle details
    "chola_type": "result.chola2x_response.data.type",
    "chola_class": "result.chola2x_response.data.class",
    "chola_model": "result.chola2x_response.data.model",
    "chola_owner": "result.chola2x_response.data.owner",
    "chola_regNo": "result.chola2x_response.data.regNo",
    "chola_engine": "result.chola2x_response.data.engine",
    "chola_status": "result.chola2x_response.data.status",
    "chola_chassis": "result.chola2x_response.data.chassis",
    "chola_regDate": "result.chola2x_response.data.regDate",
    "chola_rtoCode": "result.chola2x_response.data.rtoCode",
    "chola_bodyType": "result.chola2x_response.data.bodyType",
    "chola_nonUseTo": "result.chola2x_response.data.nonUseTo",
    "chola_puccUpto": "result.chola2x_response.data.puccUpto",
    "chola_normsType": "result.chola2x_response.data.normsType",
    "chola_wheelbase": "result.chola2x_response.data.wheelbase",
    "chola_nocDetails": "result.chola2x_response.data.nocDetails",
    "chola_nonUseFrom": "result.chola2x_response.data.nonUseFrom",
    "chola_ownerCount": "result.chola2x_response.data.ownerCount",
    "chola_permitType": "result.chola2x_response.data.permitType",
    "chola_puccNumber": "result.chola2x_response.data.puccNumber",
    "chola_rcFinancer": "result.chola2x_response.data.rcFinancer",
    "chola_statusAsOn": "result.chola2x_response.data.statusAsOn",
    "chola_isCommercial": "result.chola2x_response.data.isCommercial",
    "chola_mobileNumber": "result.chola2x_response.data.mobileNumber",
    "chola_nonUseStatus": "result.chola2x_response.data.nonUseStatus",
    "chola_permitNumber": "result.chola2x_response.data.permitNumber",
    "chola_rcExpiryDate": "result.chola2x_response.data.rcExpiryDate",
    "chola_regAuthority": "result.chola2x_response.data.regAuthority",
    "chola_rcStandardCap": "result.chola2x_response.data.rcStandardCap",
    "chola_unladenWeight": "result.chola2x_response.data.unladenWeight",
    "chola_vehicleColour": "result.chola2x_response.data.vehicleColour",
    "chola_vehicleNumber": "result.chola2x_response.data.vehicleNumber",
    "chola_presentAddress": "result.chola2x_response.data.presentAddress",
    "chola_vehicleTaxUpto": "result.chola2x_response.data.vehicleTaxUpto",
    "chola_blacklistStatus": "result.chola2x_response.data.blacklistStatus",
    "chola_ownerFatherName": "result.chola2x_response.data.ownerFatherName",
    "chola_permitIssueDate": "result.chola2x_response.data.permitIssueDate",
    "chola_permitValidFrom": "result.chola2x_response.data.permitValidFrom",
    "chola_permitValidUpto": "result.chola2x_response.data.permitValidUpto",
    "chola_vehicleCategory": "result.chola2x_response.data.vehicleCategory",
    "chola_permanentAddress": "result.chola2x_response.data.permanentAddress",
    "chola_grossVehicleWeight": "result.chola2x_response.data.grossVehicleWeight",
    "chola_nationalPermitUpto": "result.chola2x_response.data.nationalPermitUpto",
    "chola_vehicleCylindersNo": "result.chola2x_response.data.vehicleCylindersNo",
    "chola_vehicleSeatCapacity": "result.chola2x_response.data.vehicleSeatCapacity",
    "chola_nationalPermitNumber": "result.chola2x_response.data.nationalPermitNumber",
    "chola_vehicleCubicCapacity": "result.chola2x_response.data.vehicleCubicCapacity",
    "chola_vehicleInsuranceUpto": "result.chola2x_response.data.vehicleInsuranceUpto",
    "chola_nationalPermitIssuedBy": "result.chola2x_response.data.nationalPermitIssuedBy",
    "chola_vehicleSleeperCapacity": "result.chola2x_response.data.vehicleSleeperCapacity",
    "chola_vehicleManufacturerName": "result.chola2x_response.data.vehicleManufacturerName",
    "chola_vehicleStandingCapacity": "result.chola2x_response.data.vehicleStandingCapacity",
    "chola_vehicleInsuranceCompanyName": "result.chola2x_response.data.vehicleInsuranceCompanyName",
    "chola_vehicleInsurancePolicyNumber": "result.chola2x_response.data.vehicleInsurancePolicyNumber",
    "chola_vehicleManufacturingMonthYear": "result.chola2x_response.data.vehicleManufacturingMonthYear",

    # chola2x_response.data.splitPresentAddress
    "chola_present_city": "result.chola2x_response.data.splitPresentAddress.city[0]",
    "chola_present_state": "result.chola2x_response.data.splitPresentAddress.state[0][0]",
    "chola_present_state_code": "result.chola2x_response.data.splitPresentAddress.state[0][1]",
    "chola_present_country": "result.chola2x_response.data.splitPresentAddress.country[0]",
    "chola_present_pincode": "result.chola2x_response.data.splitPresentAddress.pincode",
    "chola_present_district": "result.chola2x_response.data.splitPresentAddress.district[0]",

    # chola2x_response.data.splitPermanentAddress
    "chola_permanent_city": "result.chola2x_response.data.splitPermanentAddress.city[0]",
    "chola_permanent_state": "result.chola2x_response.data.splitPermanentAddress.state[0][0]",
    "chola_permanent_state_code": "result.chola2x_response.data.splitPermanentAddress.state[0][1]",
    "chola_permanent_country": "result.chola2x_response.data.splitPermanentAddress.country[0]",
    "chola_permanent_pincode": "result.chola2x_response.data.splitPermanentAddress.pincode",
    "chola_permanent_district": "result.chola2x_response.data.splitPermanentAddress.district[0]",

    # chola2x_response.data.mappings
    "chola_signzyID": "result.chola2x_response.data.mappings.signzyID",
    "chola_variantIds": "result.chola2x_response.data.mappings.variantIds",

    # chola2x_response.data - lists
    "chola_challanDetails": "result.chola2x_response.data.challanDetails",
    "chola_blacklistDetails": "result.chola2x_response.data.blacklistDetails",
}

# ─── VALUE CONVERTER (to string) ───────────────────────────────────────────
def convert_value(value):
    """Convert any value to a string safe for Excel"""
    if value is None:
        return ""
    elif isinstance(value, bool):
        return "Yes" if value else "No"
    elif isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except:
            return str(value)
    else:
        return str(value).strip()

# ─── NESTED VALUE FETCHER ──────────────────────────────────────────────────
def get_nested(data, path):
    """Navigate nested dictionary using dot notation"""
    try:
        keys = path.replace("]", "").split(".")
        for key in keys:
            if "[" in key:
                k, idx = key.split("[")
                data = data[k][int(idx)]
            else:
                data = data.get(key, {})
        return data if data != {} else ""
    except:
        return ""

# ─── FLATTEN RESPONSE (CLEAN) ──────────────────────────────────────────────
def flatten_full(data):
    """Flatten API response to dictionary with all fields"""
    flat = {}

    for key, path in FULL_RESPONSE_KEYS.items():
        value = get_nested(data, path)
        # Convert value to string to prevent Excel errors
        flat[key] = convert_value(value)

    return flat

# ─── FETCH API ─────────────────────────────────────────────────────────────
def fetch_vehicle(vehicle_no):
    """Fetch vehicle data from API with retry logic"""
    url = f"{BASE_URL}?key={API_KEY}&type={API_TYPE}&term={vehicle_no}"

    for attempt in range(RETRIES):
        try:
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                data = r.json()

                if not data.get("success"):
                    return {
                        "status_flag": "FAILED",
                        "error_message": "API returned success=false"
                    }

                flat = flatten_full(data)
                
                # Determine status based on response
                owner = flat.get("chola_owner", "")
                reg_no = flat.get("chola_regNo", "")
                
                flat["status_flag"] = "FOUND" if (owner or reg_no) else "NOT_FOUND"
                flat["error_message"] = ""

                return flat
                
        except requests.exceptions.RequestException as e:
            if attempt < RETRIES - 1:
                time.sleep(0.5)
            else:
                return {
                    "status_flag": "ERROR",
                    "error_message": str(e)
                }

    return {
        "status_flag": "ERROR",
        "error_message": "Max retries exceeded"
    }

# ─── EXCEL BUILDER ─────────────────────────────────────────────────────────
def build_excel(rows):
    """Build Excel workbook from results with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Data"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Vehicle No", "Status", "Error Message"] + list(FULL_RESPONSE_KEYS.keys())
    ws.append(headers)

    # Apply header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25

    for idx, row in enumerate(rows, start=2):
        try:
            values = [
                convert_value(row.get("vehicle_no", "")),
                convert_value(row.get("status_flag", "")),
                convert_value(row.get("error_message", ""))
            ]
            
            # Add all other fields
            for header in FULL_RESPONSE_KEYS.keys():
                value = row.get(header, "")
                values.append(convert_value(value))
            
            ws.append(values)
        except Exception as e:
            st.warning(f"Error writing row {idx}: {str(e)}")
            continue

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── STREAMLIT UI ──────────────────────────────────────────────────────────
st.set_page_config(page_title="Vehicle Data Fetcher", layout="wide")

st.title("🚗 Vehicle RC Data Fetcher (Complete)")
st.markdown("Fetch detailed vehicle RC information with ALL fields")

# Sidebar Configuration
with st.sidebar:
    st.header("⚙️ Configuration")
    prefix = st.text_input("Vehicle Prefix", "RJ02CH", help="e.g., RJ02CH, RJ60CC")
    start = st.number_input("Start Number", 1, 9999, 1, help="Starting vehicle number")
    end = st.number_input("End Number", 1, 9999, 10, help="Ending vehicle number")
    delay = st.slider("Delay (seconds)", 0.1, 5.0, 0.5, help="Delay between API calls")

# Main UI
col1, col2, col3 = st.columns([2, 2, 1])

with col1:
    st.subheader("📋 Search Parameters")
    st.write(f"**Prefix:** `{prefix}`")
    st.write(f"**Range:** {start} to {end} (Total: {end - start + 1})")
    st.write(f"**Delay:** {delay}s between calls")

with col2:
    st.subheader("📊 Data Collection")
    st.write(f"**Fields per vehicle:** {len(FULL_RESPONSE_KEYS)}")
    st.write(f"**Excel columns:** {len(FULL_RESPONSE_KEYS) + 3}")
    st.write("**Status:** FOUND / NOT_FOUND / ERROR")

with col3:
    st.write("")
    st.write("")
    if st.button("🚀 Start Fetching", use_container_width=True, type="primary"):
        start_fetch = True
    else:
        start_fetch = False

if start_fetch:
    results = []
    progress_bar = st.progress(0)
    status_text = st.empty()
    results_container = st.container()

    total = end - start + 1

    with results_container:
        st.subheader("🔄 Fetching Results")
        results_placeholder = st.empty()
        stats_placeholder = st.empty()

    found_count = 0
    error_count = 0
    not_found_count = 0

    for idx, num in enumerate(range(start, end + 1)):
        vehicle_no = f"{prefix}{num:04d}"

        data = fetch_vehicle(vehicle_no)
        data["vehicle_no"] = vehicle_no

        results.append(data)

        # Count statuses
        status = data.get('status_flag', '')
        if status == "FOUND":
            found_count += 1
        elif status == "NOT_FOUND":
            not_found_count += 1
        else:
            error_count += 1

        # Update progress
        progress = (idx + 1) / total
        progress_bar.progress(progress)
        
        status_text.write(
            f"**Progress:** {idx + 1}/{total} | "
            f"**Current:** {vehicle_no} → {status}"
        )

        # Display result in grid
        with results_placeholder.container():
            col1, col2, col3, col4 = st.columns([2, 1, 2, 2])
            with col1:
                st.write(f"**{vehicle_no}**")
            with col2:
                status = data.get('status_flag')
                if status == "FOUND":
                    st.success(status)
                elif status == "NOT_FOUND":
                    st.warning(status)
                else:
                    st.error(status)
            with col3:
                owner = data.get('chola_owner', '')
                if owner:
                    st.write(f"👤 {owner[:30]}")
            with col4:
                if data.get('error_message'):
                    st.caption(f"⚠️ {data.get('error_message')}")

        # Update stats
        with stats_placeholder.container():
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Fetched", idx + 1)
            col2.metric("Found", found_count, delta=f"+{found_count}")
            col3.metric("Not Found", not_found_count)
            col4.metric("Errors", error_count)

        time.sleep(delay)

    # Download Section
    st.divider()
    st.subheader("📥 Download Results")

    try:
        excel = build_excel(results)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.download_button(
                "📊 Download Excel",
                excel,
                file_name=f"vehicle_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col2:
            st.metric("Total Records", len(results))

        with col3:
            st.metric("Records Found", found_count)

        with col4:
            success_rate = (found_count / len(results) * 100) if results else 0
            st.metric("Success Rate", f"{success_rate:.1f}%")

        # Show sample data
        if found_count > 0:
            st.subheader("📋 Sample Found Records")
            found_records = [r for r in results if r.get('status_flag') == 'FOUND']
            
            for record in found_records[:3]:  # Show first 3 found records
                with st.expander(f"🔍 {record.get('vehicle_no')} - {record.get('chola_owner')}"):
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.write("**Vehicle Information**")
                        st.write(f"- **Model:** {record.get('chola_model', 'N/A')}")
                        st.write(f"- **Class:** {record.get('chola_class', 'N/A')}")
                        st.write(f"- **Type:** {record.get('chola_type', 'N/A')}")
                        st.write(f"- **Variant:** {record.get('vnum_variant', 'N/A')}")
                        st.write(f"- **Color:** {record.get('chola_vehicleColour', 'N/A')}")
                    
                    with col2:
                        st.write("**Registration Details**")
                        st.write(f"- **Reg No:** {record.get('chola_regNo', 'N/A')}")
                        st.write(f"- **Reg Date:** {record.get('chola_regDate', 'N/A')}")
                        st.write(f"- **RTO:** {record.get('chola_regAuthority', 'N/A')}")
                        st.write(f"- **Status:** {record.get('chola_status', 'N/A')}")
                        st.write(f"- **RC Expiry:** {record.get('chola_rcExpiryDate', 'N/A')}")
                    
                    with col3:
                        st.write("**Insurance & PUC**")
                        st.write(f"- **Insurance Co:** {record.get('chola_vehicleInsuranceCompanyName', 'N/A')}")
                        st.write(f"- **Insurance Upto:** {record.get('chola_vehicleInsuranceUpto', 'N/A')}")
                        st.write(f"- **PUC Upto:** {record.get('chola_puccUpto', 'N/A')}")
                        st.write(f"- **Tax Upto:** {record.get('chola_vehicleTaxUpto', 'N/A')}")
                        st.write(f"- **Financer:** {record.get('chola_rcFinancer', 'N/A')}")

    except Exception as e:
        st.error(f"❌ Error creating Excel file: {str(e)}")
        st.info("Data was fetched successfully but there was an issue creating the Excel file.")

st.divider()
st.caption("🔐 Complete Vehicle RC Data Fetcher | All Fields Captured")