import time
import io
import re
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── CONSTANTS ─────────────────────────────────────────────
BASE_URL = "https://api.subhxcosmo.in/api"
API_KEY = "DEMO6"
RETRIES = 2

HEADERS = [
    "#", "Vehicle No", "Status", "Owner", "Mobile",
    "Vehicle Type", "Model", "Fuel", "Class",
    "Manufacturer", "Reg Date", "RTO",
    "Engine No", "Chassis No",
    "Insurance Expiry", "Insurance Company",
    "PUCC Upto", "Color",
    "Seat Capacity", "Cubic Capacity",
    "Owner Father", "Address"
]

# ─── FETCH FUNCTION (FIXED) ────────────────────────────────
def fetch_vehicle(vehicle_no: str) -> dict:
    url = f"{BASE_URL}?key={API_KEY}&type=vehicle&term={vehicle_no}"

    for attempt in range(RETRIES + 1):
        try:
            r = requests.get(url, timeout=15)

            if r.status_code != 200:
                return {"_status": f"HTTP {r.status_code}"}

            data = r.json()

            if not data.get("success"):
                return {"_status": "API failed"}

            result = data.get("result", {})
            d = result.get("data", {})

            if not d:
                return {"_status": "No data"}

            return {
                "vnum": result.get("vehicle", vehicle_no),

                "owner": d.get("owner", ""),
                "mobile": d.get("mobileNumber", ""),
                "vehicle_type": result.get("vehicle_type", ""),
                "status": d.get("status", ""),

                "model": d.get("model", ""),
                "fuel": d.get("type", ""),
                "class": d.get("class", ""),
                "manufacturer": d.get("vehicleManufacturerName", ""),

                "reg_date": d.get("regDate", ""),
                "rto": d.get("regAuthority", ""),

                "engine": d.get("engine", ""),
                "chassis": d.get("chassis", ""),

                "insurance": d.get("vehicleInsuranceUpto", ""),
                "insurance_company": d.get("vehicleInsuranceCompanyName", ""),

                "pucc": d.get("puccUpto", ""),
                "color": d.get("vehicleColour", ""),

                "seat": d.get("vehicleSeatCapacity", ""),
                "cc": d.get("vehicleCubicCapacity", ""),

                "father": d.get("ownerFatherName", ""),
                "address": d.get("presentAddress", "")
            }

        except Exception as e:
            if attempt == RETRIES:
                return {"_status": str(e)}
            time.sleep(0.5)

    return {"_status": "Unknown error"}

# ─── EXCEL BUILDER ─────────────────────────────────────────
def build_excel(rows):
    wb = Workbook()
    ws = wb.active

    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h)

    for r in rows:
        row = r["row_idx"]

        ws.cell(row=row, column=1, value=r["serial"])
        ws.cell(row=row, column=2, value=r["vehicle_no"])
        ws.cell(row=row, column=3, value=r["status"])
        ws.cell(row=row, column=4, value=r.get("owner"))
        ws.cell(row=row, column=5, value=r.get("mobile"))
        ws.cell(row=row, column=6, value=r.get("vehicle_type"))
        ws.cell(row=row, column=7, value=r.get("model"))
        ws.cell(row=row, column=8, value=r.get("fuel"))
        ws.cell(row=row, column=9, value=r.get("class"))
        ws.cell(row=row, column=10, value=r.get("manufacturer"))
        ws.cell(row=row, column=11, value=r.get("reg_date"))
        ws.cell(row=row, column=12, value=r.get("rto"))
        ws.cell(row=row, column=13, value=r.get("engine"))
        ws.cell(row=row, column=14, value=r.get("chassis"))
        ws.cell(row=row, column=15, value=r.get("insurance"))
        ws.cell(row=row, column=16, value=r.get("insurance_company"))
        ws.cell(row=row, column=17, value=r.get("pucc"))
        ws.cell(row=row, column=18, value=r.get("color"))
        ws.cell(row=row, column=19, value=r.get("seat"))
        ws.cell(row=row, column=20, value=r.get("cc"))
        ws.cell(row=row, column=21, value=r.get("father"))
        ws.cell(row=row, column=22, value=r.get("address"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── PREFIX PARSER ─────────────────────────────────────────
def parse_prefix(raw):
    raw = raw.strip().upper().replace(" ", "")
    m = re.match(r"^([A-Z]{2}\d{2}[A-Z]{2})(\d{1,4})?$", raw)
    if m:
        return m.group(1), int(m.group(2)) if m.group(2) else None
    return raw, None

# ─── STREAMLIT UI ──────────────────────────────────────────
st.set_page_config(page_title="Vehicle Fetcher", layout="wide")

st.title("🚗 Vehicle Data Fetcher (NEW API)")

raw_prefix = st.text_input("Enter Prefix", "RJ59CA")
prefix, auto = parse_prefix(raw_prefix)

start = st.number_input("Start", 1, 9999, auto or 1)
end = st.number_input("End", 1, 9999, 20)
delay = st.slider("Delay", 0.1, 2.0, 0.5)

if "results" not in st.session_state:
    st.session_state.results = []

if st.button("Start Fetching"):
    st.session_state.results = []

    for i, num in enumerate(range(start, end + 1), 1):
        vno = f"{prefix}{num:04d}"
        data = fetch_vehicle(vno)

        if "_status" in data:
            status = "ERROR"
        else:
            status = "FOUND"

        row = {
            "serial": i,
            "row_idx": i + 1,
            "vehicle_no": vno,
            "status": status,
            **data
        }

        st.session_state.results.append(row)
        st.write(f"{vno} → {status}")

        time.sleep(delay)

    st.success("Done!")

# ─── DOWNLOAD ──────────────────────────────────────────────
if st.session_state.results:
    buf = build_excel(st.session_state.results)

    st.download_button(
        "Download Excel",
        buf,
        file_name=f"vehicle_data_{datetime.now().strftime('%H%M%S')}.xlsx"
    )