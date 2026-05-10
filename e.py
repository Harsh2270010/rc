import time
import io
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────
BASE_URL = "https://vapi.zeltronaddy.in/v1/vehicle"
BALANCE_URL = "https://vapi.zeltronaddy.in/v1/balance"

# Response field mappings
RESPONSE_KEYS = {
    "registration_number": "regn_no",
    "owner_name": "data.owner_name",
    "father_name": "data.f_name",
    "mobile_no": "data.mobileNo",
    "address": "data.perm_address",
    "pincode": "data.p_pincode",
    "state": "data.p_state_name",
    "district": "data.p_district",
    
    "vehicle_class": "data.vh_class_desc",
    "vehicle_category": "data.catg_desc",
    "model": "data.model_name",
    "variant": "data.variant",
    "maker": "data.maker_name",
    "body_type": "data.body_type",
    
    "engine_no": "data.eng_no",
    "chassis_no": "data.chasi_no",
    "fuel_type": "data.fuel_descr",
    "cubic_capacity": "data.cubic_cap",
    "cylinders": "data.no_cyl",
    "horsepower": "data.hp",
    
    "manufacture_year": "data.manu_yr",
    "manufacture_month": "data.manu_mon",
    
    "registration_date": "data.regn_dt",
    "registration_valid_upto": "data.regn_upto",
    "fitness_upto": "data.fit_upto",
    "tax_upto": "data.tax_upto",
    "status": "data.vh_status",
    
    "rto_code": "data.rto_code",
    "rto_name": "data.off_name",
    
    "insurance_company": "data.insurance_company",
    "insurance_policy_no": "data.insurance_policy_no",
    "insurance_upto": "data.insurance_upto",
    
    "financer": "data.financer",
    "blacklist_status": "data.blacklist_status",
    "owner_type": "data.owner_cd_descr",
    
    "seating_capacity": "data.seat_cap",
    "unladen_weight": "data.unld_wt",
    "loaded_weight": "data.ld_wt",
}

# ─── NESTED VALUE FETCHER ──────────────────────────────────────────────────
def get_nested(data, path):
    """Extract nested value from dict using dot notation"""
    try:
        keys = path.split(".")
        for key in keys:
            if isinstance(data, dict):
                data = data.get(key, {})
            else:
                return ""
        return data if data else ""
    except:
        return ""

# ─── FLATTEN RESPONSE ──────────────────────────────────────────────────────
def flatten_response(api_response):
    """Flatten API response to readable format"""
    flat = {}
    
    if not api_response.get("success"):
        return None
    
    for key, path in RESPONSE_KEYS.items():
        value = get_nested(api_response, path)
        flat[key] = value if value else ""
    
    return flat

# ─── FETCH API ─────────────────────────────────────────────────────────────
def fetch_vehicle(vehicle_no, api_key, retries=2):
    """Fetch vehicle data from API"""
    headers = {"X-API-Key": api_key}
    params = {"reg": vehicle_no}
    
    for attempt in range(retries):
        try:
            response = requests.get(BASE_URL, headers=headers, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                flat = flatten_response(data)
                
                if flat:
                    return {
                        "status": "FOUND",
                        "data": flat,
                        "credits_used": data.get("_meta", {}).get("credits_used", 0),
                        "credits_remaining": data.get("_meta", {}).get("credits_remaining", 0)
                    }
                else:
                    return {"status": "NOT_FOUND"}
            
            elif response.status_code == 401:
                return {"status": "ERROR", "error": "Invalid API Key"}
            elif response.status_code == 402:
                return {"status": "ERROR", "error": "No credits remaining"}
            else:
                return {"status": "ERROR", "error": f"API Error: {response.status_code}"}
        
        except requests.exceptions.Timeout:
            if attempt == retries - 1:
                return {"status": "ERROR", "error": "Request timeout"}
            time.sleep(0.5)
        except Exception as e:
            if attempt == retries - 1:
                return {"status": "ERROR", "error": str(e)}
            time.sleep(0.5)
    
    return {"status": "ERROR", "error": "Max retries exceeded"}

# ─── CHECK BALANCE ─────────────────────────────────────────────────────────
def check_balance(api_key):
    """Check account balance"""
    headers = {"X-API-Key": api_key}
    
    try:
        response = requests.get(BALANCE_URL, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.json()
    except:
        pass
    
    return None

# ─── EXCEL BUILDER ─────────────────────────────────────────────────────────
def build_excel(results):
    """Build Excel file from results"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Data"
    
    # Headers
    headers = ["Vehicle No", "Status"] + list(RESPONSE_KEYS.keys())
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Data rows
    found_count = 0
    for result in results:
        vehicle_no = result.get("vehicle_no")
        status = result.get("status")
        
        row_data = [vehicle_no, status]
        
        if status == "FOUND":
            found_count += 1
            data = result.get("data", {})
            row_data.extend([data.get(k, "") for k in RESPONSE_KEYS.keys()])
        else:
            row_data.extend([""] * len(RESPONSE_KEYS))
        
        ws.append(row_data)
        
        # Style data row
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            if status == "FOUND":
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif status == "NOT_FOUND":
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── STREAMLIT UI ──────────────────────────────────────────────────────────
st.set_page_config(page_title="Vehicle Lookup", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    .main {
        padding: 2rem;
    }
    .stMetric {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 8px;
    }
    h1 {
        color: #1a1a1a;
        margin-bottom: 1rem;
    }
    .section-header {
        color: #366092;
        font-size: 1.1rem;
        font-weight: 600;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #e0e0e0;
    }
</style>
""", unsafe_allow_html=True)

st.title("🚗 Vehicle Lookup API")

# ─── SIDEBAR: API KEY & SETTINGS ───────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    
    api_key = st.text_input(
        "X-API-Key",
        type="password",
        help="Your API key from vapi.zeltronaddy.in"
    )
    
    if api_key and st.button("Check Balance"):
        with st.spinner("Fetching balance..."):
            balance = check_balance(api_key)
            if balance:
                st.success(f"✓ Balance check successful")
                st.json(balance)
            else:
                st.error("Failed to check balance. Verify your API key.")
    
    st.divider()
    
    lookup_mode = st.radio(
        "Lookup Mode",
        ["Single Vehicle", "Bulk Range"],
        help="Choose how to lookup vehicles"
    )

# ─── MAIN CONTENT ─────────────────────────────────────────────────────────

if not api_key:
    st.warning("⚠️ Please enter your API key in the sidebar to continue.")
    st.info("""
    **Getting Started:**
    1. Enter your X-API-Key in the sidebar
    2. Choose a lookup mode
    3. Start fetching vehicle data
    
    **Supported Format:** Indian vehicle registration numbers (e.g., OD01H1092, MH12AB1234)
    """)

else:
    # ─── MODE 1: SINGLE VEHICLE ───────────────────────────────────────────
    if lookup_mode == "Single Vehicle":
        st.markdown('<p class="section-header">🔍 Single Vehicle Lookup</p>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            vehicle_no = st.text_input(
                "Vehicle Registration Number",
                placeholder="e.g., OD01H1092",
                label_visibility="collapsed"
            )
        
        with col2:
            fetch_btn = st.button("Lookup", use_container_width=True, type="primary")
        
        if fetch_btn and vehicle_no:
            with st.spinner(f"Fetching data for {vehicle_no}..."):
                result = fetch_vehicle(vehicle_no, api_key)
            
            if result["status"] == "FOUND":
                st.success(f"✓ Vehicle found!")
                
                data = result["data"]
                
                # Owner Info
                st.markdown("**Owner Information**")
                col1, col2, col3 = st.columns(3)
                col1.metric("Owner", data.get("owner_name", "N/A"))
                col2.metric("Mobile", data.get("mobile_no", "N/A"))
                col3.metric("State", data.get("state", "N/A"))
                
                # Vehicle Info
                st.markdown("**Vehicle Details**")
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Class", data.get("vehicle_class", "N/A"))
                col2.metric("Maker", data.get("maker", "N/A")[:30])
                col3.metric("Model", data.get("model", "N/A"))
                col4.metric("Variant", data.get("variant", "N/A"))
                
                # Engine/Chassis
                st.markdown("**Engine & Chassis**")
                col1, col2, col3 = st.columns(3)
                col1.metric("Engine No", data.get("engine_no", "N/A"))
                col2.metric("Chassis No", data.get("chassis_no", "N/A"))
                col3.metric("Fuel Type", data.get("fuel_type", "N/A"))
                
                # Registration & Compliance
                st.markdown("**Registration & Compliance**")
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Reg Date", data.get("registration_date", "N/A"))
                col2.metric("Valid Upto", data.get("registration_valid_upto", "N/A"))
                col3.metric("Tax Upto", data.get("tax_upto", "N/A"))
                col4.metric("Status", data.get("status", "N/A"))
                
                # Insurance
                st.markdown("**Insurance**")
                col1, col2 = st.columns(2)
                col1.metric("Company", data.get("insurance_company", "N/A")[:40])
                col2.metric("Valid Upto", data.get("insurance_upto", "N/A"))
                
                # Full Data
                with st.expander("📋 Full Data"):
                    st.json(data)
                
                st.markdown(f"**Credits Used:** {result.get('credits_used', 0)} | **Credits Remaining:** {result.get('credits_remaining', 0)}")
            
            elif result["status"] == "NOT_FOUND":
                st.warning(f"⚠️ No vehicle record found for {vehicle_no}")
            
            else:
                st.error(f"❌ Error: {result.get('error', 'Unknown error')}")

    # ─── MODE 2: BULK RANGE ───────────────────────────────────────────────
    elif lookup_mode == "Bulk Range":
        st.markdown('<p class="section-header">📊 Bulk Range Lookup</p>', unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            prefix = st.text_input("Prefix", "RJ60CC", help="e.g., RJ60CC, OD01H")
        
        with col2:
            start = st.number_input("Start", 1, 9999, 1)
        
        with col3:
            end = st.number_input("End", 1, 9999, 10)
        
        with col4:
            delay = st.number_input("Delay (sec)", 0.1, 5.0, 0.5)
        
        fetch_bulk = st.button("Start Bulk Fetch", type="primary", use_container_width=True)
        
        if fetch_bulk:
            if start > end:
                st.error("Start number must be less than or equal to end number")
            else:
                results = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                results_container = st.container()
                
                total = end - start + 1
                
                for idx, num in enumerate(range(start, end + 1)):
                    vehicle_no = f"{prefix}{num:04d}"
                    
                    result = fetch_vehicle(vehicle_no, api_key)
                    
                    result_entry = {
                        "vehicle_no": vehicle_no,
                        "status": result["status"],
                        "data": result.get("data", {})
                    }
                    results.append(result_entry)
                    
                    # Update progress
                    progress = (idx + 1) / total
                    progress_bar.progress(progress)
                    status_text.markdown(f"**Processing:** {idx + 1}/{total} | **Status:** {result['status']}")
                    
                    time.sleep(delay)
                
                progress_bar.empty()
                status_text.empty()
                
                # Summary
                found = sum(1 for r in results if r["status"] == "FOUND")
                not_found = sum(1 for r in results if r["status"] == "NOT_FOUND")
                errors = sum(1 for r in results if r["status"] == "ERROR")
                
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total Lookups", total)
                col2.metric("Found", found, delta=None)
                col3.metric("Not Found", not_found)
                col4.metric("Errors", errors)
                
                # Results table
                st.markdown("**Detailed Results**")
                table_data = []
                for r in results:
                    table_data.append({
                        "Vehicle No": r["vehicle_no"],
                        "Status": r["status"],
                        "Owner": r["data"].get("owner_name", "") if r["status"] == "FOUND" else "",
                        "Mobile": r["data"].get("mobile_no", "") if r["status"] == "FOUND" else "",
                        "Model": r["data"].get("model", "") if r["status"] == "FOUND" else ""
                    })
                
                st.dataframe(table_data, use_container_width=True)
                
                # Excel Export
                excel_file = build_excel(results)
                st.download_button(
                    "📥 Download Excel Report",
                    excel_file,
                    file_name=f"vehicle_lookup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )