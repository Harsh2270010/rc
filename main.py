"""
RC Data Fetcher - Streamlit UI
Run: streamlit run rc_fetcher.py
"""

import time
import io
import re
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONSTANTS ───────────────────────────────────────────────────────────────
BASE_URL = "https://zero-vehiclex.vercel.app/api/rc"
API_KEY  = "UNKNOWN"
RETRIES  = 2

TOP_KEYS = [
    "rc_regn_no", "rc_status", "rc_status_as_on",
    "rc_owner_name", "rc_owner_first_name", "rc_owner_last_name", "rc_owner_sr",
    "rc_f_name", "rc_mobile_no",
    "rc_present_address", "rc_permanent_address",
    "rc_maker_desc", "rc_maker_model",
    "rc_color", "rc_fuel_desc", "rc_norms_desc",
    "rc_vh_class_desc", "rc_body_type_desc", "vehicle_category_description",
    "rc_vch_catg", "rc_rto_code", "rc_registered_at",
    "rc_regn_dt", "rc_manu_month_yr", "rc_fit_upto",
    "rc_tax_upto", "rc_insurance_comp", "rc_insurance_policy_no", "rc_insurance_upto",
    "rc_pucc_no", "rc_pucc_upto",
    "rc_chasi_no", "rc_eng_no",
    "rc_cubic_cap", "rc_no_cyl", "rc_wheelbase",
    "rc_seat_cap", "rc_sleeper_cap", "rc_stand_cap",
    "rc_unld_wt", "rc_gvw",
    "rc_financer", "financier_name_master",
    "rc_permit_type", "rc_permit_no", "rc_permit_issue_dt",
    "rc_permit_valid_from", "rc_permit_valid_upto",
    "rc_blacklist_status", "rc_noc_details", "rc_ncrb_status",
    "state_code", "pin_code", "crn", "response_timestamp",
]
PASS_KEYS = ["uid", "rc_cubic_cap", "rc_model", "rc_fuel_desc", "rc_make", "model_id", "make_id"]
HEADERS   = ["#", "Vehicle No", "Fetch Status"] + TOP_KEYS + ["pass_" + k for k in PASS_KEYS]

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def fetch_vehicle(vehicle_no: str) -> dict:
    url = f"{BASE_URL}?vehicle={vehicle_no}&key={API_KEY}"
    for attempt in range(RETRIES + 1):
        try:
            r = requests.get(url, timeout=15)
            if r.status_code == 200:
                return r.json()
            return {"_fetch_status": f"HTTP {r.status_code}"}
        except requests.RequestException as e:
            if attempt == RETRIES:
                return {"_fetch_status": f"Error: {e}"}
            time.sleep(1)
    return {"_fetch_status": "Unknown error"}


def flatten(data: dict) -> dict:
    row = {k: data.get(k, "") for k in TOP_KEYS}
    first = (data.get("pass_id_data") or [{}])[0]
    for k in PASS_KEYS:
        row["pass_" + k] = first.get(k, "")
    return row


def build_excel(rows: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "RC Data"
    ws.freeze_panes = "D2"

    hdr_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ok_fill   = PatternFill("solid", start_color="EBF5FB", end_color="EBF5FB")
    emp_fill  = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
    err_fill  = PatternFill("solid", start_color="FDECEA", end_color="FDECEA")
    body_font = Font(name="Arial", size=9)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al   = Alignment(horizontal="left", vertical="center")

    for col_idx, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 28

    for rd in rows:
        r      = rd["row_idx"]
        status = rd["status"]
        fill   = ok_fill if status == "OK" else (emp_fill if status == "EMPTY" else err_fill)

        ws.cell(row=r, column=1, value=rd["serial"])
        ws.cell(row=r, column=2, value=rd["vehicle_no"])
        ws.cell(row=r, column=3, value=status)
        for col_idx, header in enumerate(HEADERS[3:], 4):
            ws.cell(row=r, column=col_idx, value=rd.get("flat", {}).get(header, ""))

        for col_idx in range(1, len(HEADERS) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.font = body_font; c.fill = fill; c.border = border
            c.alignment = center if col_idx <= 3 else left_al

    widths = {"#": 6, "Vehicle No": 15, "Fetch Status": 13}
    for col_idx, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 20)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_prefix(raw: str):
    """
    Split user input like 'RJ02BJ' or 'RJ02BJ0001' into (prefix, maybe_start).
    Returns (prefix_upper, detected_start_or_None).
    """
    raw = raw.strip().upper().replace(" ", "")
    # If they pasted a full number like RJ02BJ0023, split off trailing digits
    m = re.match(r"^([A-Z]+\d{2}[A-Z]+)(\d{1,4})?$", raw)
    if m:
        return m.group(1), int(m.group(2)) if m.group(2) else None
    return raw, None


# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="RC Data Fetcher", page_icon="🚗", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.hero {
    background: linear-gradient(135deg, #0f2942 0%, #1a4a7a 55%, #1565c0 100%);
    border-radius: 16px;
    padding: 2rem 2.5rem 1.6rem;
    margin-bottom: 1.6rem;
    box-shadow: 0 8px 32px rgba(21,101,192,0.2);
}
.hero h1 {
    font-family: 'DM Serif Display', serif;
    color: #fff;
    font-size: 2rem;
    margin: 0 0 0.3rem;
}
.hero p { color: #90caf9; font-size: 0.92rem; margin: 0; }

.card {
    background: #f0f7ff;
    border: 1px solid #bbdefb;
    border-radius: 12px;
    padding: 0.9rem 1.2rem;
    text-align: center;
}
.card .val { font-size: 1.9rem; font-weight: 700; color: #1565c0; line-height: 1.1; }
.card .lbl { font-size: 0.72rem; color: #5c7a9c; font-weight: 600;
             text-transform: uppercase; letter-spacing: 0.6px; margin-top: 2px; }

.lbl-text { font-size: 0.76rem; font-weight: 600; color: #5c7a9c;
            text-transform: uppercase; letter-spacing: 0.6px; margin-bottom: 4px; }

.preview-box {
    background: #e8f4fd;
    border: 1px solid #90caf9;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    font-size: 0.88rem;
    color: #1a4a7a;
    font-weight: 600;
    margin-top: 0.5rem;
}

.log-box {
    background: #0d1b2a;
    border-radius: 10px;
    padding: 0.9rem 1.1rem;
    font-family: 'Courier New', monospace;
    font-size: 0.8rem;
    color: #64b5f6;
    max-height: 300px;
    overflow-y: auto;
    line-height: 1.65;
    border: 1px solid #1a3a5c;
}
.log-ok    { color: #4caf50; }
.log-empty { color: #78909c; }
.log-err   { color: #ef5350; }

.stButton > button {
    background: linear-gradient(135deg, #1565c0, #1976d2) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.6rem 1.8rem !important;
    font-weight: 600 !important;
    font-size: 0.97rem !important;
    font-family: 'DM Sans', sans-serif !important;
    box-shadow: 0 4px 14px rgba(21,101,192,0.28) !important;
}
.stButton > button:hover { opacity: 0.9 !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <h1>🚗 Vehicle RC Data Fetcher</h1>
  <p>Fetch registration certificate data for any vehicle series and export to a formatted Excel file</p>
</div>
""", unsafe_allow_html=True)

# ─── SERIES CONFIG ───────────────────────────────────────────────────────────

st.markdown("### ⚙️ Vehicle Series Configuration")

col_a, col_b = st.columns([1, 1], gap="large")

with col_a:
    st.markdown('<div class="lbl-text">Vehicle Series Prefix</div>', unsafe_allow_html=True)
    raw_prefix = st.text_input(
        "", value="RJ02UB",
        placeholder="e.g. RJ02UB or RJ02BJ or DL01AB",
        label_visibility="collapsed",
        key="prefix_input"
    )
    prefix, auto_start = parse_prefix(raw_prefix)

    # Show live preview
    preview_start = auto_start or 1
    st.markdown(
        f'<div class="preview-box">📋 Sample vehicle numbers: '
        f'<b>{prefix}{preview_start:04d}</b> → <b>{prefix}9999</b></div>',
        unsafe_allow_html=True
    )

with col_b:
    st.markdown('<div class="lbl-text">Number of digits in suffix</div>', unsafe_allow_html=True)
    digits = st.radio("", [4, 3], horizontal=True, label_visibility="collapsed",
                      help="Most RTO series use 4 digits (0001–9999). Some older series use 3.")
    st.markdown(
        f'<div class="preview-box">🔢 Format: <b>{prefix}{"0" * digits}</b> … '
        f'<b>{prefix}{"9" * digits}</b></div>',
        unsafe_allow_html=True
    )

st.markdown("---")
st.markdown("### 📊 Range & Speed")

col1, col2, col3 = st.columns(3, gap="large")
max_val = 9999 if digits == 4 else 999

with col1:
    st.markdown('<div class="lbl-text">Start Number</div>', unsafe_allow_html=True)
    start = st.number_input("", min_value=1, max_value=max_val,
                            value=auto_start or 1, step=1,
                            key="start", label_visibility="collapsed")
with col2:
    st.markdown('<div class="lbl-text">End Number</div>', unsafe_allow_html=True)
    end = st.number_input("", min_value=1, max_value=max_val,
                          value=min(100, max_val), step=1,
                          key="end", label_visibility="collapsed")
with col3:
    st.markdown('<div class="lbl-text">Delay between requests (sec)</div>', unsafe_allow_html=True)
    delay = st.slider("", min_value=0.1, max_value=2.0, value=0.3, step=0.1,
                      key="delay", label_visibility="collapsed")

# Estimate
if start <= end:
    total    = end - start + 1
    est_secs = total * (delay + 0.15)
    mins, secs = divmod(int(est_secs), 60)
    fmt      = f"{prefix}{{n:0{digits}d}}"
    first_no = fmt.format(n=start)
    last_no  = fmt.format(n=end)
    st.info(
        f"📋 **{total:,}** vehicles  ·  "
        f"**{first_no}** → **{last_no}**  ·  "
        f"Estimated time: **{mins}m {secs}s**"
    )
else:
    st.warning("⚠️ Start must be ≤ End.")

# ─── SESSION STATE ────────────────────────────────────────────────────────────
if "running"  not in st.session_state: st.session_state.running  = False
if "results"  not in st.session_state: st.session_state.results  = []
if "done"     not in st.session_state: st.session_state.done     = False
if "out_name" not in st.session_state: st.session_state.out_name = "RC_Data.xlsx"

# ─── START BUTTON ─────────────────────────────────────────────────────────────
st.markdown("")
btn_col, stop_col, _ = st.columns([1, 1, 4])
with btn_col:
    start_btn = st.button("▶ Start Fetching", disabled=st.session_state.running or start > end)
with stop_col:
    stop_btn = st.button("⏹ Stop", disabled=not st.session_state.running)

if stop_btn:
    st.session_state.running = False

# ─── FETCH LOOP ──────────────────────────────────────────────────────────────
if start_btn and not st.session_state.running and start <= end:
    st.session_state.running  = True
    st.session_state.results  = []
    st.session_state.done     = False
    st.session_state.out_name = f"{prefix}_RC_Data.xlsx"

    total = end - start + 1
    fmt   = f"{prefix}{{n:0{digits}d}}"

    st.markdown("---")
    st.markdown("### 🔄 Live Progress")

    m1, m2, m3, m4 = st.columns(4)
    with m1: ok_ph    = st.empty()
    with m2: empty_ph = st.empty()
    with m3: err_ph   = st.empty()
    with m4: prog_ph  = st.empty()

    def card(val, lbl):
        return f'<div class="card"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>'

    progress = st.progress(0)
    log_ph   = st.empty()

    ok_count = empty_count = err_count = 0
    log_lines = []

    for i, num in enumerate(range(start, end + 1), 1):
        if not st.session_state.running:
            st.warning(f"⏹ Stopped at {fmt.format(n=num)} after {i-1} vehicles.")
            break

        vehicle_no = fmt.format(n=num)
        data = fetch_vehicle(vehicle_no)
        time.sleep(delay)

        if "_fetch_status" in data:
            status = data["_fetch_status"]
            err_count += 1
            css = "log-err"
            row_data = {"serial": i, "row_idx": i + 1, "vehicle_no": vehicle_no,
                        "status": status, "flat": {}}
        elif not data.get("rc_regn_no") and not data.get("rc_eng_no"):
            status = "EMPTY"
            empty_count += 1
            css = "log-empty"
            row_data = {"serial": i, "row_idx": i + 1, "vehicle_no": vehicle_no,
                        "status": "EMPTY", "flat": flatten(data)}
        else:
            ok_count += 1
            css = "log-ok"
            owner = (data.get("rc_owner_name") or "").strip() or "—"
            model = (data.get("rc_maker_model") or "").strip() or "—"
            status = "OK"
            log_status = f"OK  |  {owner}  ·  {model}"
            row_data = {"serial": i, "row_idx": i + 1, "vehicle_no": vehicle_no,
                        "status": "OK", "flat": flatten(data)}
            log_lines.append(f'<span class="{css}">[{i:04d}/{total}] {vehicle_no}  →  {log_status}</span>')
            st.session_state.results.append(row_data)

            if i % 5 == 0 or i == total:
                ok_ph.markdown(card(ok_count, "Found"), unsafe_allow_html=True)
                empty_ph.markdown(card(empty_count, "Empty"), unsafe_allow_html=True)
                err_ph.markdown(card(err_count, "Errors"), unsafe_allow_html=True)
                prog_ph.markdown(card(f"{int(i/total*100)}%", "Progress"), unsafe_allow_html=True)
                progress.progress(i / total)
                log_ph.markdown(
                    f'<div class="log-box">{"<br>".join(reversed(log_lines[-40:]))}</div>',
                    unsafe_allow_html=True
                )
            continue

        st.session_state.results.append(row_data)
        log_lines.append(f'<span class="{css}">[{i:04d}/{total}] {vehicle_no}  →  {status}</span>')
        if len(log_lines) > 200:
            log_lines = log_lines[-200:]

        if i % 5 == 0 or i == total:
            ok_ph.markdown(card(ok_count, "Found"), unsafe_allow_html=True)
            empty_ph.markdown(card(empty_count, "Empty"), unsafe_allow_html=True)
            err_ph.markdown(card(err_count, "Errors"), unsafe_allow_html=True)
            prog_ph.markdown(card(f"{int(i/total*100)}%", "Progress"), unsafe_allow_html=True)
            progress.progress(i / total)
            log_ph.markdown(
                f'<div class="log-box">{"<br>".join(reversed(log_lines[-40:]))}</div>',
                unsafe_allow_html=True
            )

    st.session_state.running = False
    st.session_state.done    = True
    progress.progress(1.0)
    st.success(f"✅ Done! {ok_count} found · {empty_count} empty · {err_count} errors")

# ─── DOWNLOAD ─────────────────────────────────────────────────────────────────
if st.session_state.results and st.session_state.done:
    st.markdown("---")
    dl1, dl2, _ = st.columns([1, 1, 3])
    with dl1:
        with st.spinner("Building Excel…"):
            excel_buf = build_excel(st.session_state.results)
        st.download_button(
            label="⬇ Download Excel",
            data=excel_buf,
            file_name=st.session_state.out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl2:
        st.markdown(
            f"<br><small style='color:#5c7a9c'>📄 <b>{st.session_state.out_name}</b> · "
            f"{len(st.session_state.results)} rows</small>",
            unsafe_allow_html=True
        )