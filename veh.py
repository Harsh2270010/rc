import time
import io
import re
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── CONSTANTS ───────────────────────────────────────────────────────────────
BASE_URL = "https://api.subhxcosmo.in/api"
API_KEY = "DEMO6"
API_TYPE = "vehicle_num"
RETRIES = 2

# Keys we extract from the actual API response
RESPONSE_KEYS = [
    "vnum", "mobile_no", "challan_status", "challan_response",
    "vehicle_status", "vehicle_response", "cached", "attempt"
]

HEADERS_STATIC = ["#", "Vehicle No", "Status", "Mobile", "Challan Info", "Vehicle Info"]

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def fetch_vehicle(vehicle_no: str) -> dict:
    """Fetch vehicle data from the Subhx Cosmo API"""
    url = f"{BASE_URL}?key={API_KEY}&type={API_TYPE}&term={vehicle_no}"
    
    for attempt in range(RETRIES + 1):
        try:
            r = requests.get(url, timeout=15)
            if r.status_code == 200:
                data = r.json()
                
                # Top-level structure: {success, owner, result, cached, attempt}
                is_success = data.get("success", False)
                
                if not is_success:
                    return {"_status": "FAILED", "_message": "API returned success: false"}
                
                result = data.get("result", {})
                
                # Extract challan and vehicle info
                challan_info = result.get("challan_info", {})
                vehicle_info = result.get("vehicle_info", {})
                
                # Check if we have actual data
                challan_data = challan_info.get("data", "")
                vehicle_data = vehicle_info.get("data", "")
                challan_status = challan_info.get("status", False)
                vehicle_status = vehicle_info.get("status", False)
                challan_code = challan_info.get("response_code", 0)
                vehicle_code = vehicle_info.get("response_code", 0)
                
                # Return structured response
                return {
                    "vnum": result.get("vnum", vehicle_no),
                    "mobile_no": result.get("mobile_no", ""),
                    "success": is_success,
                    "owner": data.get("owner", ""),
                    "cached": data.get("cached", False),
                    "attempt": data.get("attempt", 1),
                    "challan_status": challan_status,
                    "challan_code": challan_code,
                    "challan_data": challan_data,
                    "challan_message": challan_info.get("response_message", ""),
                    "vehicle_status": vehicle_status,
                    "vehicle_code": vehicle_code,
                    "vehicle_data": vehicle_data,
                    "vehicle_message": vehicle_info.get("response_message", ""),
                    "_raw": data
                }
            else:
                return {"_status": f"HTTP {r.status_code}"}
        except requests.RequestException as e:
            if attempt == RETRIES:
                return {"_status": f"Error: {str(e)[:40]}"}
            time.sleep(0.5)
    return {"_status": "Unknown error"}


def flatten(data: dict) -> dict:
    """Flatten API response to match expected headers"""
    row = {
        "vnum": data.get("vnum", ""),
        "mobile_no": data.get("mobile_no", ""),
        "challan_status": data.get("challan_status", ""),
        "challan_response": data.get("challan_code", ""),
        "vehicle_status": data.get("vehicle_status", ""),
        "vehicle_response": data.get("vehicle_code", ""),
        "cached": data.get("cached", ""),
        "attempt": data.get("attempt", ""),
    }
    return row


def build_excel(rows: list) -> io.BytesIO:
    """Build formatted Excel workbook from results"""
    headers = HEADERS_STATIC
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Data"
    ws.freeze_panes = "D2"

    hdr_fill  = PatternFill("solid", start_color="0D1B2A", end_color="0D1B2A")
    hdr_font  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    ok_fill   = PatternFill("solid", start_color="E3F2FD", end_color="E3F2FD")
    emp_fill  = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
    err_fill  = PatternFill("solid", start_color="FFEBEE", end_color="FFEBEE")
    body_font = Font(name="Segoe UI", size=9)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Header row
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 28

    # Data rows
    for rd in rows:
        r = rd["row_idx"]
        status = rd["status"]
        fill = ok_fill if status == "FOUND" else (emp_fill if status == "NOT_FOUND" else err_fill)

        ws.cell(row=r, column=1, value=rd["serial"])
        ws.cell(row=r, column=2, value=rd["vehicle_no"])
        ws.cell(row=r, column=3, value=status)
        ws.cell(row=r, column=4, value=rd.get("mobile", ""))
        ws.cell(row=r, column=5, value=rd.get("challan", ""))
        ws.cell(row=r, column=6, value=rd.get("vehicle", ""))

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.font = body_font
            c.fill = fill
            c.border = border
            c.alignment = center if col_idx <= 3 else left_al

    # Column widths
    widths = {
        "#": 6, "Vehicle No": 16, "Status": 12,
        "Mobile": 14, "Challan Info": 14, "Vehicle Info": 14
    }
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 14)
    
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_prefix(raw: str):
    """
    Parse vehicle prefix and optional starting number.
    Handles: 'MP16CB' or 'MP16CB6745' or 'MP 16 CB 6745' or 'RJ02BJ'
    """
    raw = raw.strip().upper().replace(" ", "")
    
    # Match pattern: 2-letter state code + 2 digits + 2-letter district + optional digits
    m = re.match(r"^([A-Z]{2}\d{2}[A-Z]{2})(\d{1,4})?$", raw)
    if m:
        return m.group(1), int(m.group(2)) if m.group(2) else None
    
    # If no match, return as-is
    return raw, None


# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="Vehicle Data Fetcher v3", page_icon="🚗", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Inter:wght@400;500;600&display=swap');

html, body, [class*="css"] { 
    font-family: 'Inter', sans-serif; 
}

.hero {
    background: linear-gradient(135deg, #0d1b2a 0%, #1a3a5c 50%, #1565c0 100%);
    border-radius: 20px;
    padding: 2.5rem 3rem 2rem;
    margin-bottom: 2rem;
    box-shadow: 0 20px 60px rgba(13, 27, 42, 0.3);
    position: relative;
    overflow: hidden;
}

.hero::before {
    content: "";
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
    border-radius: 50%;
}

.hero h1 {
    font-family: 'Playfair Display', serif;
    color: #fff;
    font-size: 2.4rem;
    margin: 0 0 0.5rem;
    font-weight: 700;
    position: relative;
    z-index: 1;
}

.hero p { 
    color: #90caf9; 
    font-size: 0.95rem; 
    margin: 0;
    position: relative;
    z-index: 1;
    font-weight: 500;
}

.card {
    background: linear-gradient(135deg, #e3f2fd 0%, #f0f7ff 100%);
    border: 1px solid #90caf9;
    border-radius: 14px;
    padding: 1.1rem 1.4rem;
    text-align: center;
    transition: all 0.3s ease;
}

.card:hover {
    box-shadow: 0 8px 24px rgba(21, 101, 192, 0.15);
    transform: translateY(-2px);
}

.card .val { 
    font-size: 2.1rem; 
    font-weight: 700; 
    color: #0d47a1; 
    line-height: 1.1;
    font-family: 'Playfair Display', serif;
}

.card .lbl { 
    font-size: 0.73rem; 
    color: #5c7a9c; 
    font-weight: 600;
    text-transform: uppercase; 
    letter-spacing: 0.8px; 
    margin-top: 6px; 
}

.lbl-text { 
    font-size: 0.78rem; 
    font-weight: 600; 
    color: #1a3a5c;
    text-transform: uppercase; 
    letter-spacing: 0.8px; 
    margin-bottom: 8px; 
}

.preview-box {
    background: linear-gradient(135deg, #e8f4fd 0%, #f5fbff 100%);
    border: 2px solid #90caf9;
    border-radius: 10px;
    padding: 0.75rem 1.2rem;
    font-size: 0.9rem;
    color: #0d47a1;
    font-weight: 600;
    margin-top: 0.6rem;
    font-family: 'Courier New', monospace;
    letter-spacing: 0.5px;
}

.log-box {
    background: #0d1b2a;
    border-radius: 12px;
    padding: 1.1rem 1.3rem;
    font-family: 'Courier New', monospace;
    font-size: 0.82rem;
    color: #64b5f6;
    max-height: 350px;
    overflow-y: auto;
    line-height: 1.7;
    border: 1px solid #1a3a5c;
    box-shadow: inset 0 2px 8px rgba(0, 0, 0, 0.3);
}

.log-found { color: #4caf50; font-weight: 600; }
.log-notfound { color: #78909c; font-weight: 500; }
.log-err { color: #ef5350; font-weight: 600; }

.stButton > button {
    background: linear-gradient(135deg, #0d47a1, #1565c0) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.7rem 2rem !important;
    font-weight: 600 !important;
    font-size: 0.98rem !important;
    font-family: 'Inter', sans-serif !important;
    box-shadow: 0 6px 20px rgba(13, 71, 161, 0.28) !important;
    transition: all 0.2s ease !important;
}

.stButton > button:hover { 
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 28px rgba(13, 71, 161, 0.35) !important;
}

.stButton > button:active { transform: translateY(0) !important; }

input, select { 
    font-family: 'Inter', sans-serif !important;
    border-radius: 8px !important;
}

hr { border-color: #e0e0e0 !important; margin: 2rem 0 !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <h1>🚗 Vehicle Data Fetcher</h1>
  <p>Query vehicle data by registration number and export results to Excel</p>
</div>
""", unsafe_allow_html=True)

# ─── SERIES CONFIG ───────────────────────────────────────────────────────────

st.markdown("### ⚙️ Vehicle Series Configuration")

st.markdown('<div class="lbl-text">Vehicle Series Prefix</div>', unsafe_allow_html=True)
raw_prefix = st.text_input(
    "", value="RJ02CF",
    placeholder="e.g. MP16CB or RJ02BJ or DL01AB",
    label_visibility="collapsed",
    key="prefix_input"
)
prefix, auto_start = parse_prefix(raw_prefix)

preview_start = auto_start or 1
st.markdown(
    f'<div class="preview-box">📋 Sample: <b>{prefix}{preview_start:04d}</b> → <b>{prefix}9999</b></div>',
    unsafe_allow_html=True
)

st.markdown("---")
st.markdown("### 📊 Range & Speed")

col1, col2, col3 = st.columns(3, gap="large")
digits = 4
max_val = 9999

with col1:
    st.markdown('<div class="lbl-text">Start Number</div>', unsafe_allow_html=True)
    start = st.number_input("", min_value=1, max_value=max_val,
                            value=auto_start or 1, step=1,
                            key="start", label_visibility="collapsed")
with col2:
    st.markdown('<div class="lbl-text">End Number</div>', unsafe_allow_html=True)
    end = st.number_input("", min_value=1, max_value=max_val,
                          value=min(100, max_val), step=1,
                          key="end", label_visibility="collapsed")
with col3:
    st.markdown('<div class="lbl-text">Delay between requests (sec)</div>', unsafe_allow_html=True)
    delay = st.slider("", min_value=0.1, max_value=2.0, value=0.5, step=0.1,
                      key="delay", label_visibility="collapsed")

# Estimate
if start <= end:
    total    = end - start + 1
    est_secs = total * (delay + 0.2)
    mins, secs = divmod(int(est_secs), 60)
    fmt      = f"{prefix}{{n:04d}}"
    first_no = fmt.format(n=start)
    last_no  = fmt.format(n=end)
    st.info(
        f"📋 **{total:,}** vehicles  ·  "
        f"**{first_no}** → **{last_no}**  ·  "
        f"⏱️ Est. time: **{mins}m {secs}s**"
    )
else:
    st.warning("⚠️ Start must be ≤ End.")

# ─── SESSION STATE ────────────────────────────────────────────────────────────
if "running"  not in st.session_state: st.session_state.running  = False
if "results"  not in st.session_state: st.session_state.results  = []
if "done"     not in st.session_state: st.session_state.done     = False
if "out_name" not in st.session_state: st.session_state.out_name = "Vehicle_Data.xlsx"

# ─── START BUTTON ─────────────────────────────────────────────────────────────
st.markdown("")
btn_col, stop_col, _ = st.columns([1, 1, 4])
with btn_col:
    start_btn = st.button("▶ Start Fetching", disabled=st.session_state.running or start > end)
with stop_col:
    stop_btn = st.button("⏹ Stop", disabled=not st.session_state.running)

if stop_btn:
    st.session_state.running = False

# ─── FETCH LOOP ──────────────────────────────────────────────────────────────
if start_btn and not st.session_state.running and start <= end:
    st.session_state.running  = True
    st.session_state.results  = []
    st.session_state.done     = False
    st.session_state.out_name = f"{prefix}_Vehicle_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    total = end - start + 1
    fmt   = f"{prefix}{{n:04d}}"

    st.markdown("---")
    st.markdown("### 🔄 Live Progress")

    m1, m2, m3, m4 = st.columns(4)
    with m1: ok_ph    = st.empty()
    with m2: empty_ph = st.empty()
    with m3: err_ph   = st.empty()
    with m4: prog_ph  = st.empty()

    def card(val, lbl):
        return f'<div class="card"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>'

    progress = st.progress(0)
    log_ph   = st.empty()

    found_count = not_found_count = err_count = 0
    log_lines = []

    for i, num in enumerate(range(start, end + 1), 1):
        if not st.session_state.running:
            st.warning(f"⏹ Stopped at {fmt.format(n=num)} after {i-1} vehicles.")
            break

        vehicle_no = fmt.format(n=num)
        data = fetch_vehicle(vehicle_no)
        time.sleep(delay)

        # Check if we have actual vehicle data (mobile_no or data presence indicates record exists)
        # Response code 200 = found, 205 = not found
        if data.get("mobile_no") or data.get("vehicle_data") or data.get("challan_data"):
            found_count += 1
            status = "FOUND"
            css = "log-found"
            mobile = data.get("mobile_no", "—")
            challan_msg = "Yes" if data.get("challan_status") else "No"
            vehicle_msg = "Yes" if data.get("vehicle_status") else "No"
            log_status = f"FOUND  |  Mobile: {mobile}"
            
            row_data = {
                "serial": i,
                "row_idx": i + 1,
                "vehicle_no": vehicle_no,
                "status": "FOUND",
                "mobile": mobile,
                "challan": challan_msg,
                "vehicle": vehicle_msg,
            }
            st.session_state.results.append(row_data)
        
        elif "_status" in data:
            status = "ERROR"
            err_count += 1
            css = "log-err"
            log_status = data["_status"]
            row_data = {
                "serial": i,
                "row_idx": i + 1,
                "vehicle_no": vehicle_no,
                "status": "ERROR",
                "mobile": "",
                "challan": "",
                "vehicle": "",
            }
            st.session_state.results.append(row_data)
        
        else:
            status = "NOT_FOUND"
            not_found_count += 1
            css = "log-notfound"
            log_status = "Record not found in API"
            row_data = {
                "serial": i,
                "row_idx": i + 1,
                "vehicle_no": vehicle_no,
                "status": "NOT_FOUND",
                "mobile": "",
                "challan": "",
                "vehicle": "",
            }
            st.session_state.results.append(row_data)

        log_lines.append(f'<span class="{css}">[{i:04d}/{total}] {vehicle_no}  →  {log_status}</span>')
        if len(log_lines) > 200:
            log_lines = log_lines[-200:]

        if i % 5 == 0 or i == total:
            ok_ph.markdown(card(found_count, "Found"), unsafe_allow_html=True)
            empty_ph.markdown(card(not_found_count, "Not Found"), unsafe_allow_html=True)
            err_ph.markdown(card(err_count, "Errors"), unsafe_allow_html=True)
            prog_ph.markdown(card(f"{int(i/total*100)}%", "Progress"), unsafe_allow_html=True)
            progress.progress(i / total)
            log_ph.markdown(
                f'<div class="log-box">{"<br>".join(reversed(log_lines[-40:]))}</div>',
                unsafe_allow_html=True
            )

    st.session_state.running = False
    st.session_state.done    = True
    progress.progress(1.0)
    st.success(f"✅ Done! {found_count} found · {not_found_count} not found · {err_count} errors")

# ─── DOWNLOAD ─────────────────────────────────────────────────────────────────
if st.session_state.results and st.session_state.done:
    st.markdown("---")
    dl1, dl2, _ = st.columns([1, 1, 3])
    with dl1:
        with st.spinner("Building Excel…"):
            excel_buf = build_excel(st.session_state.results)
        st.download_button(
            label="⬇ Download Excel",
            data=excel_buf,
            file_name=st.session_state.out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl2:
        st.markdown(
            f"<br><small style='color:#5c7a9c'>📄 <b>{st.session_state.out_name}</b> · "
            f"{len(st.session_state.results)} rows</small>",
            unsafe_allow_html=True
        )
